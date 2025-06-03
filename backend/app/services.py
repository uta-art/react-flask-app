from .models import db, Todo
import csv
import io
from io import StringIO
from zoneinfo import ZoneInfo
from datetime import datetime
from openpyxl import Workbook, load_workbook

def get_paginated_todos(page=1, per_page=10):
    pagination = Todo.query.order_by(Todo.id.desc()).paginate(page=page, per_page=per_page, error_out=False)
    todos = [
        {"id": todo.id, "task": todo.task, "deadline": _to_jst(todo.deadline) if todo.deadline else None, "created_at": _to_jst(todo.created_at) if todo.created_at else None}
        for todo in pagination.items
    ]
    return {
        "todos": todos,
        "total": pagination.total,
        "page": page,
        "per_page": per_page,
        "pages": pagination.pages
    }

def add_todo(task, deadline=None):
    if not task:
        return False, "Task is required"
    # 期限の文字列をdatetimeオブジェクトに変換
    deadline_dt = None
    if deadline:
        try:
            deadline_dt = datetime.fromisoformat(deadline.replace('Z', '+00:00'))
        except ValueError:
            return False, "Invalid deadline format"
    todo = Todo(task=task, deadline=deadline_dt)
    db.session.add(todo)
    db.session.commit()
    return True, None

def delete_todo(todo_id):
    todo = Todo.query.get(todo_id)
    if not todo:
        return False, "Not found"
    db.session.delete(todo)
    db.session.commit()
    return True, None

def update_todo(todo_id, new_task, new_deadline=None):
    todo = Todo.query.get(todo_id)
    if not todo:
        return False, "Not found"
    todo.task = new_task
    # 期限の更新
    if new_deadline:
        try:
            todo.deadline = datetime.fromisoformat(new_deadline.replace('Z', '+00:00'))
        except ValueError:
            return False, "Invalid deadline format"
    else:
        todo.deadline = None
    db.session.commit()
    return True, None

def import_todos_from_csv(file_storage):
    stream = io.StringIO(file_storage.stream.read().decode('utf-8'))
    reader = csv.reader(stream)
    count = 0
    for row in reader:
        if not row or not row[0].strip():
            continue
        task = row[0].strip()
        deadline = None
        # CSV に期限列がある場合の処理（2列目）
        if len(row) > 1 and row[1].strip():
            try:
                deadline = datetime.fromisoformat(row[1].strip().replace('Z', '+00:00'))
            except ValueError:
                pass  # 期限の形式が不正な場合は None のまま
        todo = Todo(task=task, deadline=deadline)
        db.session.add(todo)
        count += 1
    db.session.commit()
    return count

def import_todos_from_excel(file_storage):
    try:
        # Excelファイルを読み込み
        # read_only=True: メモリ効率化のため読み取り専用モードで開く
        workbook = load_workbook(file_storage, read_only=True)

        # 最初のワークシート（シート）を取得
        # .active: アクティブなワークシートを取得するプロパティ
        worksheet = workbook.active

        count = 0  # インポート件数カウンタ

        # iter_rows(): 行を順次取得するイテレータ
        # min_row=2: 2行目から開始（1行目はヘッダーとして除外）
        # values_only=True: セルのフォーマット情報を除外し、値のみ取得
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # row: タプル形式 例: ('タスク名', '2025-06-01 14:00', None, ...)

            # 空行チェック（rowがNoneまたは最初の要素が空）
            if not row or not row[0]:
                continue  # この行をスキップして次の行へ

            # 1列目（インデックス0）からタスク名を取得
            # str()で文字列に変換してstrip()で前後の空白を除去
            task = str(row[0]).strip() if row[0] else None
            if not task:  # タスク名が空の場合はスキップ
                continue

            deadline = None  # 期限の初期値

            # 2列目（インデックス1）に期限がある場合の処理
            # len(row) > 1: 2列目以降が存在するかチェック
            # row[1]: 2列目の値をチェック
            if len(row) > 1 and row[1]:
                try:
                    # isinstance(): オブジェクトの型チェック
                    if isinstance(row[1], datetime):
                        # 既にdatetimeオブジェクトの場合（Excelの日付セル）
                        deadline = row[1]
                    else:
                        # 文字列の場合の処理
                        deadline_str = str(row[1]).strip()
                        if deadline_str:
                            # ISO形式の日時文字列をdatetimeオブジェクトに変換
                            # replace('Z', '+00:00'): UTC時刻表記の正規化
                            deadline = datetime.fromisoformat(deadline_str.replace('Z', '+00:00'))

                except (ValueError, TypeError):
                    # ValueError: 日時フォーマットが不正
                    # TypeError: 型変換エラー
                    # エラーが発生した場合はdeadlineをNoneのまま（スキップ）
                    pass

            # Todoオブジェクトを作成
            # SQLAlchemyのモデルインスタンス作成
            todo = Todo(task=task, deadline=deadline)

            # データベースセッションに追加（まだコミットしない）
            db.session.add(todo)
            count += 1  # カウンタをインクリメント

        # 全ての変更をデータベースにコミット（実際に保存）
        db.session.commit()

        # Excelファイルを閉じてメモリを解放
        workbook.close()

        return count  # インポート件数を返す

    except Exception as e:
        db.session.rollback()

        raise Exception(f"Excelファイルの処理中にエラーが発生しました: {str(e)}")

def export_todos_csv():
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['id', 'task', 'deadline', 'created_at'])
    todos = Todo.query.order_by(Todo.id.desc()).all()
    for todo in todos:
        writer.writerow([
            todo.id,
            todo.task,
            todo.deadline.isoformat() if todo.deadline else '',
            todo.created_at
        ])
    output.seek(0)
    return output

def bulk_delete_todos(ids):
    try:
        Todo.query.filter(Todo.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        return True, None
    except Exception as e:
        db.session.rollback()
        return False, str(e)

def _to_jst(dt):
    if dt is None:
        return None
    # UTC→JSTへ変換
    return dt.astimezone(ZoneInfo("Asia/Tokyo")).strftime('%Y/%m/%d %H:%M')